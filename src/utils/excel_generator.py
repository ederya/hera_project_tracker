import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from src.database import get_all_tasks_with_people, get_all_people
import io

def generate_excel(lang='EN'):
    """Generates an Excel file with multiple sheets, one for each person."""
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet created by openpyxl
    default_sheet = wb.active
    
    # Language setup
    if lang == 'TR':
        period_mapping = {
            1: "Bu Hafta",
            2: "Gelecek Hafta",
            3: "Bu Ay İçinde",
            4: "Beklemede"
        }
        headers = ["Aksiyon Adı", "Durum", "Hedef Dönem", "Kesin Tarih", "İlgili Kişiler", "Sonuç"]
        open_text = "Açık"
        closed_text = "Kapalı"
        no_data_text = "Görev bulunamadı"
    else:
        period_mapping = {
            1: "This Week",
            2: "Next Week",
            3: "This Month",
            4: "On Hold"
        }
        headers = ["Action", "Status", "Target Period", "Exact Date", "Related People", "Result"]
        open_text = "Open"
        closed_text = "Closed"
        no_data_text = "No tasks available"
    
    # Get all tasks joined with people
    data = get_all_tasks_with_people()
    all_people = get_all_people()
    
    # Group tasks by person
    people_tasks = {}
    for person in all_people:
        people_tasks[person['name']] = []
        
    for row in data:
        person_name = row['person_name']
        people_tasks[person_name].append(row)
        
    if not people_tasks:
        # If no data, just rename the default sheet and add headers
        default_sheet.title = "No Data"
        default_sheet.append([no_data_text])
    else:
        # We will remove the default sheet after creating our own
        is_first_sheet = True
        
        from src.database import get_tasks_waiting_on_person
        
        for person in all_people:
            person_name = person['name']
            person_id = person['id']
            tasks = people_tasks[person_name]
            
            # Create a sheet for the person
            sheet = wb.create_sheet(title=person_name)
            
            # If it's the first person, we can remove the default sheet
            if is_first_sheet:
                wb.remove(default_sheet)
                is_first_sheet = False
            
            # Define Borders
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            # Add big title for person's name
            sheet.append([person_name.upper()])
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
            title_cell = sheet.cell(row=1, column=1)
            title_cell.font = Font(bold=True, size=16, color="000000")
            title_cell.fill = PatternFill("solid", fgColor="DDDDDD")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for col in range(1, 7):
                sheet.cell(row=1, column=col).border = thin_border
            
            # Define Headers
            sheet.append(headers)
            
            # Style the header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4682B4") # Steel Blue
            
            for cell in sheet[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            
            # Append tasks
            if not tasks:
                sheet.append([no_data_text])
                sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
            else:
                for task in tasks:
                    target_period = period_mapping.get(task['target_period_id'], "Unknown")
                    exact_date = task['exact_date'] if task['is_exact_date_active'] else "-"
                    result_translated = open_text if task['result'] == 'Open' else closed_text
                    
                    action_name = task['action_name']
                    related_people = []
                    if task.get('collaborators'):
                        for collab in task['collaborators']:
                            if collab['is_completed']:
                                if lang == 'TR':
                                    warning = f"[{collab['name']} tamamladı]"
                                else:
                                    warning = f"[{collab['name']} completed]"
                            else:
                                if lang == 'TR':
                                    warning = f"[Bekleniyor: {collab['name']}]"
                                else:
                                    warning = f"[Waiting for: {collab['name']}]"
                            
                            if collab.get('waiting_reason'):
                                warning += f" - \"{collab['waiting_reason']}\""
                            related_people.append(warning)
                            
                    related_people_str = "\n".join(related_people) if related_people else "-"
                    
                    sheet.append([
                        action_name,
                        task['status'],
                        target_period,
                        exact_date,
                        related_people_str,
                        result_translated
                    ])
                
            # Adjust column widths to standard sizes (wider to fit screen)
            sheet.column_dimensions['A'].width = 50  # Action
            sheet.column_dimensions['B'].width = 50  # Status
            sheet.column_dimensions['C'].width = 20  # Target Period
            sheet.column_dimensions['D'].width = 20  # Exact Date
            sheet.column_dimensions['E'].width = 30  # Related People
            sheet.column_dimensions['F'].width = 15  # Result
            
            # Check for waiting tasks
            waiting_tasks = get_tasks_waiting_on_person(person_id)
            if waiting_tasks:
                sheet.append([]) # empty row
                sub_title = "Ondan Beklenen Diğer Görevler" if lang == 'TR' else "Tasks Waiting On Them"
                sheet.append([sub_title])
                sub_row = sheet.max_row
                sheet.merge_cells(start_row=sub_row, start_column=1, end_row=sub_row, end_column=6)
                sub_cell = sheet.cell(row=sub_row, column=1)
                sub_cell.font = Font(bold=True, size=11, color="FFFFFF")
                sub_cell.fill = PatternFill("solid", fgColor="FFA500") # Orange
                sub_cell.alignment = Alignment(horizontal="center", vertical="center")
                for col in range(1, 7):
                    sheet.cell(row=sub_row, column=col).border = thin_border
                
                sheet.append(headers)
                h_row = sheet.max_row
                for col, cell in enumerate(sheet[h_row], start=1):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border
                    
                for task in waiting_tasks:
                    target_period = period_mapping.get(task['target_period_id'], "Unknown")
                    exact_date = task['exact_date'] if task['is_exact_date_active'] else "-"
                    result_translated = open_text if task['result'] == 'Open' else closed_text
                    
                    action_name = task['action_name']
                    related_people = []
                    
                    # Show who is the owner
                    owner_text = f"[İşin Sahibi: {task['owner_name']}]" if lang == 'TR' else f"[Owner: {task['owner_name']}]"
                    related_people.append(owner_text)
                    
                    if task.get('waiting_reason'):
                        related_people.append(f"\"{task['waiting_reason']}\"")
                        
                    related_people_str = "\n".join(related_people)
                        
                    sheet.append([
                        action_name,
                        task['status'],
                        target_period,
                        exact_date,
                        related_people_str,
                        result_translated
                    ])
            
            # Apply wrap text and border to all data cells
            wrap_alignment = Alignment(wrap_text=True, vertical="top")
            for row in sheet.iter_rows(min_row=3):
                for cell in row:
                    cell.alignment = wrap_alignment
                    if cell.value != None:
                        cell.border = thin_border

    # Save to a memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
